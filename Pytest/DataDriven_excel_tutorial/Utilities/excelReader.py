import os
import openpyxl

def get_data(path, sheet_name):
    final_list = []

    base_dir = os.path.dirname(os.path.dirname(__file__))
    file_path = os.path.join(base_dir, path)

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    for r in range(2, sheet.max_row + 1):
        row_list = []
        for c in range(1, sheet.max_column + 1):
            row_list.append(sheet.cell(r, c).value)
        final_list.append(row_list)

    return final_list