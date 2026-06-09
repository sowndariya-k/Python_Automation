import os
import openpyxl

def get_data(relative_path, sheet_name):
    BASE_DIR = os.getcwd()   # Jenkins workspace root during run
    file_path = os.path.join(BASE_DIR, relative_path)

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found at: {file_path}")

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    data = []
    for r in range(2, sheet.max_row + 1):
        row = []
        for c in range(1, sheet.max_column + 1):
            row.append(sheet.cell(r, c).value)
        data.append(tuple(row))

    return data